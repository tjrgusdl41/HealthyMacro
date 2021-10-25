import sys

import PyQt5.QtCore
from PyQt5.QtWidgets import *
from PyQt5 import uic, QtCore
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from heal import health_pa
from openpyxl import Workbook, load_workbook
import numpy as np

# UI파일 연결
# 단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form_class = uic.loadUiType("health.ui")[0]

wb = None
chrows = []


def doWb():
    global wb
    try:
        wb = load_workbook("db.xlsx")
    except Exception as e:
        wb = Workbook("db.xlsx")
    return wb


class MyQTableWidgetItemCheckBox(QTableWidgetItem):
    """
        checkbox widget 과 같은 cell 에 item 으로 들어감.
        checkbox 값 변화에 따라, 사용자정의 data를 기준으로 정렬 기능 구현함.
    """

    def __init__(self):
        super().__init__()
        self.setData(Qt.UserRole, 0)

    def __lt__(self, other):
        # print(type(self.data(Qt.UserRole)))
        return self.data(Qt.UserRole) < other.data(Qt.UserRole)

    def my_setdata(self, value):
        # print("my setdata ", value)
        self.setData(Qt.UserRole, value)
        self.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
        # print("row ", self.row())


class MyCheckBox(QCheckBox):
    def __init__(self, item):
        """ :param item: QTableWidgetItem instance """
        super().__init__()
        self.item = item
        self.mycheckvalue = 0  # 0 --> unchecked, 2 --> checked
        self.stateChanged.connect(self.__checkbox_change)
        self.stateChanged.connect(self.item.my_setdata)  # checked 여부로 정렬을 하기위한 data 저장

    def __checkbox_change(self, checkvalue):
        global chrows
        # print("myclass...check change... ", checkvalue)
        self.mycheckvalue = checkvalue
        print("checkbox row= ", self.get_row())
        try:
            if int(self.get_row()) in chrows:
                chrows.remove(self.get_row())
            else:
                chrows.append(self.get_row())
        except Exception as e:
            print(e)
        print(chrows)

    def get_row(self):
        return self.item.row()


# 화면을 띄우는데 사용되는 Class 선언
class WindowClass(QMainWindow, form_class):
    def __init__(self):
        super().__init__()
        self.setupUi(self)

        self.tablesetdate()

        """
        ---------------------------------------------
        이 부분에 시그널을 입력해야 합니다.
        시그널이 작동할 때 실행될 기능은 보통 이 클래스의 멤버함수로 작성합니다.
        ---------------------------------------------
        """

        # 버튼에 기능을 연결하는 코드
        self.btn1.clicked.connect(self.delData)
        self.btn2.clicked.connect(self.addData)
        self.btn1_2.clicked.connect(self.exeData)

    def _cellclicked(self, row, col):
        print("_cellclicked... ", row, col)

    # btn_1이 눌리면 작동할 함수
    def delData(self):
        wb = doWb()
        global chrows
        try:
            ws = wb.active
            np1 = np.array(chrows)
            for item in np1:
                ws.delete_rows(item + 1)
                np1[item < np1] -= 1
            wb.save("db.xlsx")
        except Exception as e:
            print(e)
        finally:
            wb.close()
        self.tablesetdate()
        chrows.clear()

    # btn_2가 눌리면 작동할 함수
    def addData(self):
        name = self.tb_name.text()
        bitrh = self.tb_birth.text()
        pw = self.tb_pw.text()

        try:
            if name == "" or bitrh == "" or pw == "":
                print("실패")
                return
        except Exception as e:
            print(e)
        wb = doWb()
        try:
            ws = wb.active
            ws.append((name, bitrh, pw))
            wb.save("db.xlsx")

        except Exception as e:
            print(e)
        finally:
            wb.close()
        self.tablesetdate()
        chrows.clear()

    # btn_2가 눌리면 작동할 함수
    def exeData(self):
        wb = doWb()
        try:
            ws = wb.active
            np1 = np.array(chrows)
            for idx in chrows:
                name = self.table.item(idx, 1).text()
                birth = self.table.item(idx, 2).text()
                pw = self.table.item(idx, 3).text()
                health_pa(name, birth, pw)
        except Exception as e:
            print(e)
        finally:
            wb.close()


    def __checkbox_change(self, checkvalue):  # print("check change... ", checkvalue)
        chbox = self.sender()  # signal을 보낸 MyCheckBox instance
        print("checkbox sender row = ", chbox.get_row())

    def tablesetdate(self):
        wb = doWb()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["", "이름", "생년월일", "비밀번호"])
        data = []

        ws = wb.active
        cnt = 0
        for i in ws.iter_rows():
            data.append((i[0].value, i[1].value, i[2].value))
            cnt += 1
        wb.close()
        self.table.setRowCount(cnt)

        for idx, (hname, birth, passwd) in enumerate(data):
            # 사용자정의 item 과 checkbox widget 을, 동일한 cell 에 넣어서 , 추후 정렬 가능하게 한다.
            item = MyQTableWidgetItemCheckBox()
            self.table.setItem(idx, 0, item)
            chbox = MyCheckBox(item)  # print(chbox.sizeHint())
            chbox.setStyleSheet("margin-left:15px;")
            self.table.setCellWidget(idx, 0, chbox)
            chbox.stateChanged.connect(self.__checkbox_change)  # sender() 확인용 예..
            # 이름
            item = QTableWidgetItem(hname)
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.table.setItem(idx, 1, item)
            # 나이
            item = QTableWidgetItem(birth)
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            self.table.setItem(idx, 2, item)  # 숫자를 기준으로 정렬하기 위함. -- default 는 '문자'임.
            # 비밀번호
            item = QTableWidgetItem()
            item.setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
            item.setData(Qt.DisplayRole, passwd)
            self.table.setItem(idx, 3, item)
            # self.table.setItem(idx, 4, QTableWidgetItem(vol))

        self.table.setSortingEnabled(False)  # 정렬기능
        self.table.resizeRowsToContents()
        self.table.resizeColumnsToContents()  # 이것만으로는 checkbox 컬럼은 잘 조절안됨.
        self.table.setColumnWidth(0, 15)  # checkbox 컬럼 폭 강제 조절.
        self.table.setColumnWidth(1, 145)
        self.table.setColumnWidth(2, 155)
        self.table.setColumnWidth(3, 145)
        self.table.cellClicked.connect(self._cellclicked)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec_()
